import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import io
import re
import requests
import base64
import openpyxl

# ─────────────────────────────────────────────────────────────────
# GITHUB CONFIGURATION
# ─────────────────────────────────────────────────────────────────
GITHUB_USER = "LinoVation1312"
GITHUB_REPO = "database"
BRANCH = "master"

if "GITHUB_TOKEN" in st.secrets:
    TOKEN = st.secrets["GITHUB_TOKEN"]
else:
    TOKEN = None

headers = {"Authorization": f"token {TOKEN}", "Accept": "application/vnd.github.v3+json"} if TOKEN else {}
FILENAME_REGEX = r"^database[-_\s]?v.*\.xlsx$"

# ─────────────────────────────────────────────────────────────────
# GITHUB SEARCH / READ / WRITE FUNCTIONS
# ─────────────────────────────────────────────────────────────────
@st.cache_data(ttl=60)
def find_and_download_current_file():
    contents_url = f"https://api.github.com/repos/{GITHUB_USER}/{GITHUB_REPO}/contents/"
    try:
        res = requests.get(contents_url, headers=headers, params={"t": pd.Timestamp.now().timestamp()})
        if res.status_code == 200:
            files = res.json()
            for f in files:
                if re.match(FILENAME_REGEX, f["name"].lower()):
                    file_res = requests.get(f["download_url"], headers=headers, params={"t": pd.Timestamp.now().timestamp()})
                    if file_res.status_code == 200:
                        return f["name"], file_res.content
        return None, None
    except Exception as e:
        return None, None

def upload_new_excel_to_github(new_filename, file_bytes, commit_message="Database update via Streamlit"):
    if not TOKEN:
        st.error("❌ The GITHUB_TOKEN is missing from the application secrets.")
        return False

    contents_url = f"https://api.github.com/repos/{GITHUB_USER}/{GITHUB_REPO}/contents/"
    try:
        existing_sha = None
        old_filename_to_delete = None

        res = requests.get(contents_url, headers=headers, params={"t": pd.Timestamp.now().timestamp()})
        if res.status_code == 200:
            for f in res.json():
                if re.match(FILENAME_REGEX, f["name"].lower()):
                    if f["name"].lower() == new_filename.lower():
                        existing_sha = f["sha"]
                        new_filename = f["name"]
                    else:
                        old_filename_to_delete = f["name"]
                        old_file_sha = f["sha"]

        upload_url = f"https://api.github.com/repos/{GITHUB_USER}/{GITHUB_REPO}/contents/{new_filename}"
        content_b64 = base64.b64encode(file_bytes).decode("utf-8")
        
        put_data = {
            "message": commit_message,
            "content": content_b64,
            "branch": BRANCH
        }
        if existing_sha: put_data["sha"] = existing_sha

        put_res = requests.put(upload_url, headers=headers, json=put_data)
        
        if put_res.status_code not in [200, 201]:
            st.error(f"❌ GitHub push failed (Code {put_res.status_code}): {put_res.text}")
            return False

        if old_filename_to_delete:
            delete_url = f"https://api.github.com/repos/{GITHUB_USER}/{GITHUB_REPO}/contents/{old_filename_to_delete}"
            del_data = {
                "message": f"Cleanup old version {old_filename_to_delete} after upgrade",
                "sha": old_file_sha,
                "branch": BRANCH
            }
            requests.delete(delete_url, headers=headers, json=del_data)

        return True

    except Exception as e:
        st.error(f"Error during GitHub synchronization: {e}")
        return False

# ─────────────────────────────────────────────────────────────────
# PAGE CONFIGURATION
# ─────────────────────────────────────────────────────────────────
st.set_page_config(page_title="DATABASE", page_icon="🔊", layout="wide")

st.markdown("""
<style>
    [data-testid="stSidebar"] { background-color: #f8fafc; border-right: 1px solid #e2e8f0; }
    [data-testid="stSidebar"] * { color: #0f172a !important; }
    .stMetric { background-color: #f1f5f9; padding: 15px; border-radius: 10px; border: 1px solid #e2e8f0; width: fit-content; }
    h1 { color: #1e40af !important; font-weight: 800 !important; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────
# HELPER FUNCTIONS (Parsing logic)
# ─────────────────────────────────────────────────────────────────
def norm_cols(cols):
    return (cols.str.strip().str.lower().str.replace(r'[^\w\s]', '', regex=True).str.replace(r'\s+', '_', regex=True))

MATERIAL_MAP = [(r'\bGlass\s*[Ff]iber\b', 'GF'), (r'\bPANox\b|\bPANOX\b', 'PANox'), (r'\bPES\b', 'PES'), (r'\bPET\b', 'PET'), (r'\bPP\b', 'PP')]
STN_PATTERN = r'^(E\d+|REF\s+\S.*)$'

def parse_materials(description: str) -> str:
    if not isinstance(description, str): return "?"
    hits = [(m.start(), label) for pattern, label in MATERIAL_MAP for m in re.finditer(pattern, description, re.IGNORECASE)]
    hits.sort(key=lambda x: x[0])
    seen, found = set(), []
    for _, label in hits:
        if label not in seen: seen.add(label); found.append(label)
    return "+".join(found) if found else "?"

def parse_airgap(text: str):
    if not isinstance(text, str): return None
    m = re.search(r'(\d+)\s*mm\s*air\s*?gap|air\s*?gap\s*(\d+)\s*mm', text, re.IGNORECASE)
    return f"{m.group(1) or m.group(2)} mm" if m else None

def is_ref(row: pd.Series) -> bool:
    return str(row.get("stn", "")).upper().startswith("REF ")

def is_composite(row: pd.Series) -> bool:
    if is_ref(row): return False
    stn  = str(row.get("stn", "")).upper()
    desc = str(row.get("detailed_description", ""))
    if "COMP" in stn: return True
    if "+" in desc:
        parts = desc.split("+")
        if len(parts) >= 2:
            mat_in = lambda t: any(re.search(pat, t, re.IGNORECASE) for pat, _ in MATERIAL_MAP)
            if mat_in(parts[0]) and mat_in("+".join(parts[1:])): return True
    return False

def parse_composite_layers(description: str, mass_col_val):
    if not isinstance(description, str): return None, None
    parts = description.split("+")
    if len(parts) < 2: return None, None

    def layer_info(text, fallback_mass=None):
        mats = parse_materials(text)
        mass_m = re.search(r'(\d[\d,]*)\s*(?:gsm|gm|g/m²|g/m2)', text, re.IGNORECASE)
        mass = (mass_m.group(1).replace(",", "") if mass_m else (str(int(float(fallback_mass))) if pd.notna(fallback_mass) else "?"))
        thick_m = re.search(r'(\d+(?:[.,]\d+)?)\s*mm', text, re.IGNORECASE)
        thick = thick_m.group(1) if thick_m else None
        label = f"{mats}  {mass} gsm"
        if thick: label += f"  {thick} mm"
        return label

    l1 = layer_info(parts[0].strip(), fallback_mass=mass_col_val)
    l2 = layer_info("+".join(parts[1:]).strip())
    return l1, l2

def build_curve_label(row: pd.Series, mass_col: str) -> str:
    stn  = str(row.get("stn", "?")).strip()
    desc = str(row.get("detailed_description", ""))
    mass = row.get(mass_col)
    thick = row.get("thickness_mm")

    if is_ref(row): return f"★ {stn}"

    thick_str = f"{thick} mm" if pd.notna(thick) else "? mm"
    airgap    = parse_airgap(str(row.get("material_orientation", ""))) or parse_airgap(desc)
    ag_str    = f" | AG {airgap}" if airgap else ""

    if is_composite(row):
        l1, l2 = parse_composite_layers(desc, mass)
        if l1 and l2: return f"⊕ {stn} | [{l1}] + [{l2}] | {thick_str}{ag_str}"
        mat      = parse_materials(desc)
        mass_str = f"{int(float(mass))} gsm" if pd.notna(mass) else "? gsm"
        return f"⊕ {stn} | {mat} | {mass_str} | {thick_str}{ag_str}"

    mat      = parse_materials(desc)
    mass_str = f"{int(float(mass))} gsm" if pd.notna(mass) else "? gsm"
    return f"{stn} | {mat} | {mass_str} | {thick_str}{ag_str}"

# ─────────────────────────────────────────────────────────────────
# DATA LOADING & CLEANING
# ─────────────────────────────────────────────────────────────────
def load_data(file_bytes: bytes):
    buf = io.BytesIO(file_bytes)
    xf  = pd.ExcelFile(buf, engine="openpyxl")

    gnrl_sheet = next((s for s in xf.sheet_names if s.strip().upper().startswith("GNRL")), None)
    abs_sheet  = next((s for s in xf.sheet_names if s.strip().upper() == "ABSORPTION"), None)
    stl_sheet  = next((s for s in xf.sheet_names if s.strip().upper() == "STL"), None)

    if not gnrl_sheet: return None, None, None, None
    
    # Store exact sheet names for appending later
    sheet_names = {"GNRL": gnrl_sheet, "ABS": abs_sheet, "STL": stl_sheet}

    raw = pd.read_excel(buf, sheet_name=gnrl_sheet, engine="openpyxl", header=None)
    header_row = next((i for i, r in raw.iterrows() if any("sample" in str(v).lower() or "stn" in str(v).lower() for v in r if pd.notna(v))), None)
    if header_row is None: return None, None, None, None

    gnrl = pd.read_excel(buf, sheet_name=gnrl_sheet, engine="openpyxl", header=header_row)
    gnrl.columns = norm_cols(gnrl.columns)
    gnrl = gnrl.dropna(how="all")

    stn_cols = [c for c in gnrl.columns if "stn" in c or "sample" in c]
    if not stn_cols: return None, None, None, None

    short_col = next((c for c in stn_cols if gnrl[c].dropna().astype(str).str.strip().str.match(r'^E\d+').mean() > 0.4), stn_cols[-1])
    gnrl = gnrl.rename(columns={short_col: "stn"})
    gnrl["stn"] = gnrl["stn"].astype(str).str.strip().str.upper()
    gnrl = gnrl[gnrl["stn"].str.match(STN_PATTERN, na=False)]

    mass_col = next((c for c in gnrl.columns if "surface_mass" in c), None)
    if mass_col: gnrl[mass_col] = pd.to_numeric(gnrl[mass_col], errors="coerce")
    gnrl["thickness_mm"] = pd.to_numeric(gnrl.get("thickness_mm"), errors="coerce")

    gnrl["is_ref"]       = gnrl.apply(is_ref, axis=1)
    gnrl["is_composite"] = gnrl.apply(is_composite, axis=1)

    def process_data_sheet(sheet_name):
        if not sheet_name: return pd.DataFrame()
        data = pd.read_excel(buf, sheet_name=sheet_name, engine="openpyxl")
        data.columns = norm_cols(data.columns)

        if "stn" not in data.columns:
            stn_data_col = next((c for c in data.columns if "stn" in c or "sample" in c), None)
            if stn_data_col: data = data.rename(columns={stn_data_col: "stn"})
            else: return pd.DataFrame()

        for c in data.columns:
            if "stl" in c or "alpha_cabin_stl" in c: data = data.rename(columns={c: "alpha_cabin_stl"})
            elif "alpha_cabin" in c: data = data.rename(columns={c: "alpha_cabin"})
            elif "alpha_kundt" in c: data = data.rename(columns={c: "alpha_kundt"})
            elif "frequency"   in c: data = data.rename(columns={c: "frequency"})

        data["stn"] = (data["stn"].astype(str).replace({"nan": pd.NA, "None": pd.NA, "": pd.NA}).ffill().str.strip().str.upper())
        data = data[data["stn"].str.match(STN_PATTERN, na=False)]

        for col in ["frequency", "alpha_cabin", "alpha_kundt", "alpha_cabin_stl"]:
            if col in data.columns: data[col] = pd.to_numeric(data[col], errors="coerce")

        merged = data.merge(gnrl, on="stn", how="left")
        merged["is_ref"]       = merged["is_ref"].fillna(False)
        merged["is_composite"] = merged["is_composite"].fillna(False)
        merged["curve_label"]  = merged.apply(lambda r: build_curve_label(r, mass_col), axis=1)
        return merged

    df_abs = process_data_sheet(abs_sheet)
    df_stl = process_data_sheet(stl_sheet)

    return df_abs, df_stl, gnrl, sheet_names

# ─────────────────────────────────────────────────────────────────
# MAIN UI & DYNAMIC LOADING
# ─────────────────────────────────────────────────────────────────
st.title("🔊 DATABASE")

current_filename, excel_data = find_and_download_current_file()

with st.sidebar.expander("🔄 GitHub Administration", expanded=False):
    uploaded_file = st.file_uploader("Upload a new database file", type=["xlsx"])
    if uploaded_file:
        file_bytes = uploaded_file.read()
        filename_uploaded = uploaded_file.name
        if re.match(FILENAME_REGEX, filename_uploaded.lower()):
            if st.button("🚀 Overwrite & Publish Version"):
                with st.spinner("Uploading to GitHub..."):
                    if upload_new_excel_to_github(filename_uploaded, file_bytes):
                        st.success("✅ Upload successful!")
                        st.cache_data.clear()
                        st.rerun()

if excel_data is None:
    st.error("❌ No valid database file was found on GitHub.")
    st.stop()

st.caption(f"📂 Active database loaded from GitHub: `{current_filename}`")

df_abs, df_stl, df_gnrl, sheet_names = load_data(excel_data)
if df_gnrl is None: st.stop()

mass_col = next((c for c in df_gnrl.columns if "surface_mass" in c), "surface_mass_gsm")

# ─────────────────────────────────────────────────────────────────
# SIDEBAR FILTERS
# ─────────────────────────────────────────────────────────────────
st.sidebar.header("📁 Data Category")
data_options = []
if not df_abs.empty: data_options.append("Absorption")
if not df_stl.empty: data_options.append("STL")
data_type = st.sidebar.radio("Select Category to Analyze", data_options) if data_options else None

if data_type == "Absorption":
    df = df_abs
    available_methods = [c for c in ["alpha_cabin", "alpha_kundt"] if c in df.columns]
    abs_type = st.sidebar.radio("Measurement Method", available_methods) if available_methods else None
elif data_type == "STL":
    df = df_stl
    available_methods = [c for c in ["alpha_cabin_stl"] if c in df.columns]
    abs_type = st.sidebar.radio("Measurement Method", available_methods) if available_methods else "alpha_cabin_stl"
else:
    df = df_gnrl
    abs_type = None

st.sidebar.header("🎛️ Global Filters")

trim_sel = (st.sidebar.multiselect("Trim Level", sorted(df["trim_level"].dropna().unique())) if "trim_level" in df.columns else [])
sup_sel  = (st.sidebar.multiselect("Material Supplier", sorted(df["material_supplier"].dropna().unique())) if "material_supplier" in df.columns else [])

m_min = float(df[mass_col].min(skipna=True) or 0)
m_max = float(df[mass_col].max(skipna=True) or 100)
mass_range = st.sidebar.slider("Surface Mass", m_min, m_max, (m_min, m_max))

fdf_samples = df[~df["is_ref"]].copy()
if trim_sel: fdf_samples = fdf_samples[fdf_samples["trim_level"].isin(trim_sel)]
if sup_sel:  fdf_samples = fdf_samples[fdf_samples["material_supplier"].isin(sup_sel)]
fdf_samples = fdf_samples[fdf_samples[mass_col].between(*mass_range) | fdf_samples[mass_col].isna()]

fdf = pd.concat([fdf_samples, df[df["is_ref"]]], ignore_index=True)

st.sidebar.markdown("---")
available_labels = sorted(fdf["curve_label"].dropna().unique().tolist())
select_all = st.sidebar.checkbox("Select All Samples", value=False)
selected_labels = st.sidebar.multiselect(f"Select Samples ({len(available_labels)} available)", available_labels, default=available_labels if select_all else [])

all_active_labels = selected_labels
plot_data = fdf[fdf["curve_label"].isin(all_active_labels)]

# ─────────────────────────────────────────────────────────────────
# TABS
# ─────────────────────────────────────────────────────────────────
tab1, tab2, tab3 = st.tabs(["📈 Interactive Plot", "🗃️ Raw Data & Exports", "➕ Ajouter une matière"])

with tab1:
    if not all_active_labels:
        st.info("👈 Select at least one sample from the sidebar to generate the charts.")
    elif abs_type:
        import plotly.io as pio
        fig = go.Figure()
        for label in all_active_labels:
            sub = plot_data[plot_data["curve_label"] == label].dropna(subset=["frequency", abs_type])
            if sub.empty: continue
            sub = sub.groupby("frequency", as_index=False)[abs_type].mean().sort_values("frequency")
            
            is_ref_curve = label.startswith("★")
            fig.add_trace(go.Scatter(
                x=sub["frequency"], y=sub[abs_type], mode="lines+markers", name=label,
                line=dict(dash="dash" if is_ref_curve else "solid", width=3 if is_ref_curve else 2)
            ))

        fig.update_layout(
            title=f"{data_type} ({abs_type})", xaxis=dict(type="log", title="Frequency (Hz)"), yaxis=dict(title="Value"),
            hovermode="x unified", legend=dict(orientation="h", y=-0.2)
        )
        st.plotly_chart(fig, use_container_width=True)

with tab2:
    st.markdown("### 📥 Source File Download")
    st.download_button("🟢 Download Current Excel", data=excel_data, file_name=current_filename)
    if not plot_data.empty and abs_type:
        st.markdown("### 📊 Formatted Data")
        pivot_data = plot_data.groupby(['frequency', 'curve_label'], as_index=False)[abs_type].mean()
        wide_df = pivot_data.pivot(index="frequency", columns="curve_label", values=abs_type).reset_index()
        st.dataframe(wide_df, hide_index=True)

# ─────────────────────────────────────────────────────────────────
# TAB 3 : AJOUT D'ÉCHANTILLON (CRUD)
# ─────────────────────────────────────────────────────────────────
with tab3:
    st.header("➕ Ajouter un nouvel échantillon")
    st.markdown("Complétez ce formulaire pour injecter de nouvelles données directement dans le fichier Excel source.")
    
    # --- 1. Détermination du prochain STN ---
    existing_stns = df_gnrl['stn'].dropna().astype(str).tolist()
    e_nums = [int(re.search(r'\d+', s).group()) for s in existing_stns if s.startswith('E') and re.search(r'\d+', s)]
    next_id = max(e_nums) + 1 if e_nums else 1
    new_stn = f"E{next_id:04d}"
    
    st.subheader(f"Nouvel Identifiant : **{new_stn}**")
    
    # --- 2. Helper pour les dropdowns "Autre" ---
    def get_options(col_name):
        opts = [""]
        if col_name in df_gnrl.columns:
            opts += sorted([str(x) for x in df_gnrl[col_name].dropna().unique() if str(x).strip() != ""])
        opts.append("➕ Autre (nouveau)")
        return opts

    # --- 3. Formulaire (sans st.form pour permettre le re-run dynamique des dropdowns) ---
    colA, colB = st.columns(2)
    
    with colA:
        # Fournisseur
        sup_opts = get_options("material_supplier")
        sel_sup = st.selectbox("Fournisseur (Material Supplier)", sup_opts)
        final_sup = st.text_input("Saisir le nouveau fournisseur", key="new_sup") if sel_sup == "➕ Autre (nouveau)" else sel_sup
        
        # Famille
        fam_opts = get_options("material_family")
        sel_fam = st.selectbox("Famille (Material Family)", fam_opts)
        final_fam = st.text_input("Saisir la nouvelle famille", key="new_fam") if sel_fam == "➕ Autre (nouveau)" else sel_fam
        
        # Trim Level
        trim_opts = get_options("trim_level")
        sel_trim = st.selectbox("Trim Level", trim_opts)
        final_trim = st.text_input("Saisir le nouveau Trim Level", key="new_trim") if sel_trim == "➕ Autre (nouveau)" else sel_trim

    with colB:
        # Masse Surfacique
        final_mass = st.number_input("Masse Surfacique (g/m²)", min_value=0.0, step=10.0)
        
        # Épaisseur
        thick_opts = get_options("thickness_mm")
        sel_thick = st.selectbox("Épaisseur (mm)", thick_opts)
        final_thick = st.number_input("Saisir la nouvelle épaisseur (mm)", min_value=0.0, step=0.5, key="new_thick") if sel_thick == "➕ Autre (nouveau)" else sel_thick
        
        # Description détaillée
        final_desc = st.text_input("Description Détaillée (ex: 1 layer PET)")

    st.markdown("---")
    st.subheader("📊 Données Acoustiques")
    
    # Grille de saisie des fréquences par défaut
    default_freqs = [200, 250, 315, 400, 500, 630, 800, 1000, 1250, 1600, 2000, 2500, 3150, 4000, 5000, 6300, 8000, 10000]
    input_df = pd.DataFrame({
        "Frequency": default_freqs,
        "Alpha Cabin": [None] * len(default_freqs),
        "Alpha Kundt": [None] * len(default_freqs),
        "STL (dB)": [None] * len(default_freqs)
    })
    
    st.info("💡 Saisissez vos valeurs ci-dessous. Laissez vide si la donnée n'est pas disponible.")
    edited_data = st.data_editor(input_df, num_rows="dynamic", use_container_width=True)

    # --- 4. Logique de sauvegarde (Écriture Excel via Openpyxl) ---
    if st.button("💾 Sauvegarder l'échantillon et mettre à jour le serveur", type="primary"):
        with st.spinner("Ouverture et modification du fichier Excel..."):
            try:
                # Chargement du classeur Excel en mémoire
                wb = openpyxl.load_workbook(io.BytesIO(excel_data))
                
                # Ajout dans la feuille GNRL
                if sheet_names["GNRL"] in wb.sheetnames:
                    ws_gnrl = wb[sheet_names["GNRL"]]
                    # Trouver la colonne d'en-tête (on cherche la ligne d'en-tête)
                    header_row_idx = None
                    for row_idx, row in enumerate(ws_gnrl.iter_rows(values_only=True), 1):
                        if any(str(cell).lower().find("sample") != -1 for cell in row if cell):
                            header_row_idx = row_idx
                            headers_gnrl = [str(cell).strip().lower() for cell in row]
                            break
                    
                    if header_row_idx:
                        new_row = [""] * len(headers_gnrl)
                        
                        def fill_col(col_keyword, val):
                            for i, h in enumerate(headers_gnrl):
                                if h and col_keyword in h: new_row[i] = val; break
                        
                        # Remplissage des données
                        fill_col("sample number", new_stn)
                        fill_col("stn", new_stn)
                        fill_col("supplier", final_sup)
                        fill_col("family", final_fam)
                        fill_col("trim", final_trim)
                        fill_col("mass", final_mass)
                        fill_col("thickness", final_thick)
                        fill_col("description", final_desc)
                        fill_col("date", pd.Timestamp.now().strftime("%Y-%m-%d"))
                        
                        ws_gnrl.append(new_row)

                # Ajout dans la feuille ABSORPTION
                if sheet_names["ABS"] in wb.sheetnames:
                    ws_abs = wb[sheet_names["ABS"]]
                    for _, row in edited_data.iterrows():
                        if pd.notna(row["Alpha Cabin"]) or pd.notna(row["Alpha Kundt"]):
                            ws_abs.append([new_stn, row["Frequency"], row["Alpha Cabin"], row["Alpha Kundt"]])
                            
                # Ajout dans la feuille STL
                if sheet_names["STL"] in wb.sheetnames:
                    ws_stl = wb[sheet_names["STL"]]
                    for _, row in edited_data.iterrows():
                        if pd.notna(row["STL (dB)"]):
                            # Format classique STL : STN, Freq, Valeur
                            ws_stl.append([new_stn, row["Frequency"], row["STL (dB)"]])

                # Sauvegarde du nouveau fichier en mémoire
                output_buffer = io.BytesIO()
                wb.save(output_buffer)
                new_excel_bytes = output_buffer.getvalue()

                # Envoi vers GitHub
                commit_message = f"Ajout du nouvel échantillon {new_stn} via l'interface Streamlit"
                if upload_new_excel_to_github(current_filename, new_excel_bytes, commit_message):
                    st.success(f"✅ L'échantillon {new_stn} a été ajouté avec succès ! La page va s'actualiser.")
                    import time
                    time.sleep(2)
                    st.cache_data.clear()
                    st.rerun()

            except Exception as e:
                st.error(f"❌ Erreur lors de la modification de l'Excel : {e}")
